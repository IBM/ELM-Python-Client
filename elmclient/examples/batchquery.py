##
## © Copyright 2021- IBM Inc. All rights reserved
# SPDX-License-Identifier: MIT
##


#
# This code has two use-cases:
# 1. Where you want to run a batch of queries, with each saving results in CSV - specify the TestId this is added to make filename))
# 2. When you want to run a series of queries and check the data retrieved is the same as the last time your ran it
#    i.e. as a regression tester - with the queries specified in the spreadsheet and fully cached data used as a mock
#    server  you don't even need to have the server attached
#
# This is a simple regression tester which compares current results with saved results from a previous run on the same server.
# The server doesn't always have to be attached to re-run tests because the test mode uses 'forever' http caching to save everything.
# Although if the OSLC query causes a request for a previously-unfetched response from the server then things will of course fail.
# But that doesn't seem too unreasonable, as everything gets 'forever' cached including login response :-)
#

import argparse
import re
import sys
import time

import openpyxl as XL

import elmclient.examples.oslcquery as querymain

# this maps important column headings to commandline option+value - these columns must all be present in the worksheet - if the value in a cell is None then no prefix is put in the commandline
# if a heading isn't in this list it is ignored
# the sequence of these determines the sequence they appear in the oslc query commandline
# (otherwise they would have been in alphabetical order)
xlstoargs={
    'Appstring':       '-A'
    ,'Project':         '-p'
    ,'Component':       '-C'
    ,'Configuration':   '-F'
    ,'GC Config':       '-G'
    ,'GC Project':      '-E'
    ,'ResourceType':    '-r'
    ,'Query':            '-q'
    ,'Select':          '-s'
    ,'Searchterms':     '-f'
    ,"Orderby":         '-o'
    ,'Null':            '-n'
    ,'Value':           '-v'
    ,'Value1':          '-v'
    ,'OutputFile':      '-O'
    ,'TypeSystemReport': '--typesystemreport'
    ,'Browser':         '-B'
    ,'Creds0':          '-0'
    ,'Creds1':          '-1'
    ,'Creds2':          '-2'
    ,'Creds3':          '-3'
    ,'Creds4':          '-4'
    ,'NResults':        '--nresults'
    ,'User':            '-U'
    ,'Password':        '-P'
    ,'JazzURL':         '-J'
    ,'Logging':         '-L'
}

# turn a list of options into a Windows cmd-style quoted string (means this works on Windows only!)
# option strings are NOT already wrapped in quotes!
# first it doubles " in the string, then if space or " in the string it is wrapped in " "
def argstocmd(args):
    newargs = []
    for arg in args:
        if '"' in arg:
            arg = arg.replace( '"','""')
        if ' ' in arg or '"' in arg:
            arg = f'"{arg}"'
        newargs.append(arg)
    return " ".join(newargs)


def do_tests(inputargs=None):
    inputargs = inputargs or sys.argv[1:]

    # setup argparse
    parser = argparse.ArgumentParser(description="Perform OSLC query on a Jazz application, with results output to CSV (and other) formats - use -h to get some basic help")

    parser.add_argument('spreadsheet', help='Name of the xlsx spreadsheet with tests')
    parser.add_argument('-d', '--dryrun', action="store_true", help="Dry run - show commandline but don't run the OSLC Query")
    parser.add_argument('-f', '--stoponfail', action="store_true", help="Stop at first failure")
    parser.add_argument('-g', '--group', default=None, help="Comma-separated list of regex pattern to match groups to be run, in the worksheet Group column")
    parser.add_argument('-j', '--just', default=None, help="Comma-separated list of tests to run, matching the TestId column in the worksheet")
    parser.add_argument('-L', '--loglevel', default=None, help="Set logging level - default is None - choose from INFO/DEBUG/ERROR")
    parser.add_argument('-r', '--reps', default=1, type=int, help="Number of times to repeat the selected tests (must be >=1")
    parser.add_argument('-s', '--save', action="store_true", help="UNFINISHED Retrieve and save query results forever (used to save reference for -t testing")
    parser.add_argument('-t', '--test', action="store_true", help="UNFINISHED Retrieve data and do comparison to test that results match saved results from -s")
    parser.add_argument('-w', '--sheetname', default=None, help='Name of the worksheet with tests (if not specified the workbook must only have one worksheet, which is used)')
    parser.add_argument('-W', '--cachecontrol', action='count', default=0, help="Used once -W erases cache for the first test then continues with caching enabled. Used twice -WW wipes cache and disables caching.")

    args = parser.parse_args(inputargs)

    if args.reps<1:
        raise Exception( f"Reps must be >=1" )

    justtests = [j.strip() for j in args.just.split(",")] if args.just else []
    
    wb = XL.load_workbook(filename=args.spreadsheet,data_only=True)

    wss=wb.sheetnames
    if args.sheetname:
        tests = wb[args.sheetname]
    else:
        if len( wss ) > 1:
            raise Exception( "Worksheet not specified but spreadsheet file includes more than one sheet!" )
        print( f"Using worksheet {wss[0]}" )
        tests = wb[wss[0]]

    # first scan the headings on row 1 to get the column numbers for the columns we want to use
    # turn the worksheet content into a list of dictionaries using the column headings as keys
    colheadings = []
    for col in range(1,50):
        thiscolheading = tests.cell(column=col, row=1).value
        # first empty heading terminates the table
        if not thiscolheading:
            break
        colheadings.append(thiscolheading)
        # now retrieve data to list of dictionaries, one per row
    rows = []
    for rownum in range(2, 2000):
        row = {}
        for i,col in enumerate(colheadings):
            row[col]=tests.cell(column=i+1, row=rownum).value
        rows.append(row)
    wb.close()

    # now go down the rows executing the specified test
    npassed = 0
    nfailed = 0
    failedtests = []
    firstquery = True
    for rep in range(args.reps):
        for n,row in enumerate(rows):
            testnumber = row['TestId']
            if not testnumber:
                continue
            if row['Disable'] and row['Disable'].startswith('#'):
                continue
            if args.group:
                if not row['Group']:
                    continue
                rowgroups = [j.strip() for j in row['Group'].split(",")]
                regexes = [j.strip() for j in args.group.split(",")]
                if not any([re.match(regex,group) for regex in regexes for group in rowgroups]):
                    continue
            if justtests and str(testnumber) not in justtests:
                continue
                
            print( f"=====================================================================\n{testnumber=} {row.get('Description','')}" )
            exceptionexpected = True if row['ExceptionExpected'] else False
            csvname = "test_"+str(testnumber)+".csv"
            queryargs=[]
            for k,v in xlstoargs.items():
                if k not in colheadings:
                    raise Exception( f"Heading {k} not present in spreadsheet!" )
                cellvalue=row[k]
                if cellvalue is not None:
                    if v:
                        # if there's an option
                        cellvalue=str(row[k]).strip()
                        # check for options where the value starts with - - these have to be specified using -o=value
                        if cellvalue.startswith("-"):
                            # use -o=value
                            queryargs.append( f"{v}={cellvalue}" )
                        else:
                            # use -o value
                            queryargs.append(v)
                            queryargs.append(cellvalue)
                    else:
                        queryargs.append(str(cellvalue).strip())
            if args.save:
                queryargs.extend(['-0','-O',csvname])
            if args.test:
                queryargs.extend(['-0','-2',csvname])
            if args.loglevel and "-L" not in queryargs:
                queryargs.extend(['-L',args.loglevel])

            # handle cache control passing on to oslcquery
            if firstquery:
                # if this is first query run and we have to wipe cache:
                if args.cachecontrol==1:
                    queryargs.extend( [ "-W" ] )
                elif args.cachecontrol==2:
                    queryargs.extend( [ "-WW" ] )
                firstquery = False
            elif args.cachecontrol==2:
                queryargs.extend( [ "-WW" ] )

            # run it
            try:
                if args.dryrun:
                    print( f"Dry-run: query commandline is: oslcquery {argstocmd(queryargs)}" )
                    result = 0
                else:
                    print( f"Query commandline is: oslcquery {argstocmd(queryargs)}" )
                    result = querymain.do_oslc_query(queryargs)
                exceptionhappened = False
            except Exception as e:
                print( e )
                result = 1
                exceptionhappened = True
    #            if not exceptionexpected:
    #                raise
            if (result != 0 and not exceptionexpected) or (result == 0 and exceptionexpected):
                    nfailed += 1
                    failedtests.append(str(testnumber))

                    print( f" TEST {testnumber} FAILED!!!!!!!!!!!!!!!!!!!!!\n" )
                    if args.stoponfail:
                        print( f"Stopping after first failure, {rep} repetitions" )
                        return
            else:
                print( f"Test {testnumber} passed!" )
                npassed += 1

    if not args.dryrun:
        print( f"\nPassed {npassed} Failed {nfailed}" )
        if failedtests:
            faileds = '\n   '+'\n   '.join(failedtests)
            print( f"Failed tests:{faileds}" ) 
    else:
        print( f"Dry run completed" )

def main():
    runstarttime = time.perf_counter()
    do_tests(sys.argv[1:])
    elapsedsecs = time.perf_counter() - runstarttime
    print( f"Runtime was {int(elapsedsecs/60)}m {int(elapsedsecs%60):02d}s" )

if __name__ == '__main__':
    main()

